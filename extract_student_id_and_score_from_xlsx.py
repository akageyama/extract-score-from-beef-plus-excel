#!/usr/bin/env python3

'''
  2024.02.01

  BEEF+では、ある授業科目の全小テスト結果をまとめてダウンロード
  することができない。小テストを実施した回数だけ手で「答案ダウンロード」
  ボタンをクリックするする必要がある。そうしてダウンロードされた
  ファイルは「問題」と「解答」という名前の2枚のシートから構成
  されるエクセルファイルである。

  これは、そのエクセルファイルの「解答」シートから学籍番号と氏名、
  そして小テストの点数（一番右のカラム）を抽出するPythonスクリプトである。
'''

import sys
import openpyxl

args = sys.argv
target_excel_file_name = args[1]

wb = openpyxl.load_workbook(target_excel_file_name)
sheet = wb['解答']

num_of_studetns_challenged_to_this_quiz = sheet.max_row + 1
max_column_of_this_sheet = sheet.max_column

for i in range(2, num_of_studetns_challenged_to_this_quiz):
    student_name = sheet.cell(row=i,column=4).value 
    student_id = sheet.cell(row=i,column=2).value
    student_score = sheet.cell(row=i,column=max_column_of_this_sheet).value
    print(student_id, ',', student_name, ',', student_score)
